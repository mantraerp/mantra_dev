# Copyright (c) 2025, Foram Shah and contributors
# For license information, please see license.txt

import base64
import frappe
import json
import re
import openpyxl
from frappe.utils import get_first_day, get_last_day
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta


def execute(filters=None):
	columns = get_columns()
	data = get_data(filters)
	return columns, data


def get_columns():
	return [
		{"label": "Employee Code", "fieldname": "employee_code", "fieldtype": "Link", "options": "Employee", "width": 200},
		{"label": "Employee Name", "fieldname": "employee_name", "fieldtype": "Data", "width": 300},
		{"label": "Designation", "fieldname": "designation", "fieldtype": "Data", "width": 200},
		{"label": "Posting Date", "fieldname": "posting_date", "fieldtype": "Date", "width": 150},
		{"label": "No. Of Days Worked", "fieldname": "no_working_days", "fieldtype": "Data", "width": 150},
		{"label": "Overtime", "fieldname": "overtime", "fieldtype": "Data", "width": 150},
		{"label": "Basic Wages", "fieldname": "basic_wages", "fieldtype": "Currency", "options": "Currency", "width": 180},
		{"label": "HRA", "fieldname": "hra", "fieldtype": "Currency", "options": "Currency", "width": 180},
		{"label": "Other cash payments", "fieldname": "other_cash_payments", "fieldtype": "Currency", "options": "Currency", "width": 150},
		{"label": "Total Gross", "fieldname": "total_gross", "fieldtype": "Currency", "options": "Currency", "width": 180},
		{"label": "PF", "fieldname": "pf", "fieldtype": "Currency", "options": "Currency", "width": 150},
		{"label": "ESIC/ WC", "fieldname": "esic_wc", "fieldtype": "Currency", "options": "Currency", "width": 150},
		{"label": "PT", "fieldname": "pt", "fieldtype": "Currency", "options": "Currency", "width": 150},
		{"label": "Net Amount Paid", "fieldname": "net_amount_paid", "fieldtype": "Currency", "options": "Currency", "width": 180},
		{"label": "Place Of Payment", "fieldname": "place_of_payment", "fieldtype": "Data", "width": 150},
	]


def get_data(filters):
	from_date = get_first_day(f"{filters.year}-{filters.month}-01")
	to_date = get_last_day(from_date)

	conditions = """
		ss.docstatus = 1
	"""

	if filters.employee_list:
		filters.employee = filters.employee + [value.strip() for value in re.split(r'[,\n]', filters['employee_list']) if value.strip()]

	if filters.employee:
		if len(filters.employee) > 1:
			conditions += f"AND ss.employee IN {tuple(filters.employee)}"
		else:
			conditions += f"AND ss.employee = '{filters.employee[0]}'"

	query = f"""
		SELECT
			ss.name,
			ss.employee as employee_code,
			ss.employee_name,
			em.designation,
			ss.total_working_days as no_working_days,
			ss.gross_pay as total_gross,
			ss.rounded_total as net_amount_paid,
			ss.posting_date,
			CASE 
				WHEN '{filters.based_on_location}' = 'Company HO' THEN 'Ahmedabad'
				ELSE em.branch
			END as place_of_payment
		FROM 
			`tabSalary Slip` as ss
		LEFT JOIN
			`tabEmployee` as em ON em.name = ss.employee
		WHERE
			ss.start_date >= %(from_date)s AND 
			ss.end_date <= %(to_date)s AND 
			{conditions}
	"""
	
	salary_slips = frappe.db.sql(query, {'from_date': from_date, 'to_date': to_date}, as_dict=True)
	for salary in salary_slips:
		if basic_amount := frappe.db.get_value("Salary Detail", {'salary_component': 'Basic', 'parent': salary.name, 'parentfield': 'earnings'}, 'amount'):
			salary['basic_wages'] = basic_amount

		if hra_amount := frappe.db.get_value("Salary Detail", {'salary_component': 'House Rent Allowance', 'parent': salary.name, 'parentfield': 'earnings'}, 'amount'):
			salary['hra'] = hra_amount

		if pf_amount := frappe.db.get_value("Salary Detail", {'salary_component': "Employee's Contribution to PF", 'parent': salary.name, 'parentfield': 'deductions'}, 'amount'):
			salary['pf'] = pf_amount

		if pt_amount := frappe.db.get_value("Salary Detail", {'salary_component': 'Professional Tax', 'parent': salary.name, 'parentfield': 'deductions'}, 'amount'):
			salary['pt'] = pt_amount

		if esic_amount := frappe.db.get_value("Salary Detail", {'salary_component': "Employee's Contribution to ESIC", 'parent': salary.name, 'parentfield': 'deductions'}, 'amount'):
			salary['esic_wc'] = esic_amount

	return salary_slips


@frappe.whitelist()
def download_excel(report_data, month, year):
	report_data = json.loads(report_data)
	month = int(month)
	year = int(year)
	first_day = datetime(year, month, 1)
	next_month = first_day.replace(day=28) + timedelta(days=4)
	last_day = next_month - timedelta(days=next_month.day)

	month_date = datetime(2025, month, 1)
	full_month_name = month_date.strftime('%B')
	short_month_name = month_date.strftime('%b')

	# Create a new Excel workbook and sheet
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"FORM XVII-{full_month_name} {year}"

	# Define styles
	bold_font = Font(name='Calibri', size=11, bold=True)
	center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

	# Write Header Section
	column_dimensions_list = [
		('A', 8), ('B', 30), ('C', 20), ('D', 24), ('E', 10), ('F', 12), ('G', 10), ('H', 10), ('I', 9), ('J', 8), ('K', 10), 
		('L', 10), ('M', 8), ('N', 8), ('O', 10), ('P', 12), ('Q', 12), ('R', 18), ('S', 16), ('T', 10), ('U', 8), ('V', 10),
	]

	for col_name, value in column_dimensions_list:
		ws.column_dimensions[col_name].width = value

	ws.row_dimensions[1].height = 64
	ws.row_dimensions[2].height = 49
	ws.row_dimensions[3].height = 24
	ws.row_dimensions[4].height = 40
	ws.row_dimensions[5].height = 90

	cell_details = [
		('A1:R1', f"FORM XVII\n[See Rule 78 (1) (a) (i)]\nRegister of Wages- {short_month_name} {year}"),
		('A2:D2', "Name and address of Contractor"),
		('E2:I2', "Mantra Softech India Pvt Ltd. 203, Shapath Hexa, Opp. Gujarat High Court, S G Highway, Ahmedabad-380060"),
		('J2:L3', "Name and address of Establishment in/under which contract is carried on"),
		('M2:R3', ""),
		('A3:B3', "Nature and work Location"),
		('C3', None),
		('D3:I3', None),
		('A4:D4', "Wage period: Monthly                                         From"),
		('E4:F4', first_day),
		('G4', "TO"),
		('H4:I4', last_day),
		('J4:L4', "Name and address of Principal Employer"),
		('A5', "Serial Number"),
		('B5', "Name of workman"),
		('C5', "Serial No, in the register of workmen"),
		('D5', "Designation / Nature of work done"),
		('E5', "No. of days worked"),
		('F5', "Units of work done"),
		('G5', "Daily rate of wages / piece rate"),
		('H5', "Basic wages"),
		('I5', "HRA"),
		('J5', "Overtime"),
		('K5', "Other cash payments (nature of payment to be indicated)"),
		('L5', "Total-gross"),
		('M5', "PF"),
		('N5', "ESIC/ WC"),
		('O5', "PT"),
		('P5', "Net amount paid"),
		('Q5', "Time and date of payment"),
		('R5', "Place of payment")
	]

	for col, header in cell_details:
		if ':' in col:
			ws.merge_cells(col)
		col = col.split(':')[0]
		ws[col].value = header
		ws[col].font = bold_font
		ws[col].border = thin_border
		ws[col].alignment = center_align

		# Apply yellow fill conditionally
		# if col in ['E5', 'H5', 'L5', 'M5', 'N5', 'O5']:
		# 	ws[col].fill = yellow_fill

		if col in ['E4', 'H4']:
			ws[col].number_format = 'd-mmm-yy'

	row_count = 6
	total_dict = {
		'no_working_days': 0,
		'basic_wages': 0,
		'hra': 0,
		'total_gross': 0,
		'pf': 0,
		'esic_wc': 0,
		'pt': 0,
		'net_amount_paid': 0
	}

	if report_data[-1]['employee_code'] == 'Total':
		report_data = report_data[:-1]

	for index, data in enumerate(report_data):
		ws[f'A{row_count}'] = index + 1
		ws[f'B{row_count}'] = data.get('employee_name')
		ws[f'C{row_count}'] = data.get('employee_code')
		ws[f'D{row_count}'] = data.get('designation')
		ws[f'E{row_count}'] = data.get('no_working_days')
		ws[f'H{row_count}'] = data.get('basic_wages')
		ws[f'I{row_count}'] = data.get('hra')
		ws[f'L{row_count}'] = data.get('total_gross')
		ws[f'M{row_count}'] = data.get('pf')
		ws[f'N{row_count}'] = data.get('esic_wc')
		ws[f'O{row_count}'] = data.get('pt')
		ws[f'P{row_count}'] = data.get('net_amount_paid')
		ws[f'Q{row_count}'] = data.get('posting_date')
		ws[f'R{row_count}'] = data.get('place_of_payment')

		for key in total_dict:
			total_dict[key] += data.get(key, 0)

		row_count += 1
	
	ws.merge_cells(f'A{row_count}:D{row_count}')
	for col in range(ord('A'), ord('R') + 1):
		ws[f'{chr(col)}{row_count}'].font = bold_font

	ws[f'A{row_count}'] = 'Total'
	ws[f'E{row_count}'] = total_dict.get('no_working_days')
	ws[f'H{row_count}'] = total_dict.get('basic_wages')
	ws[f'I{row_count}'] = total_dict.get('hra')
	ws[f'L{row_count}'] = total_dict.get('total_gross')
	ws[f'M{row_count}'] = total_dict.get('pf')
	ws[f'N{row_count}'] = total_dict.get('esic_wc')
	ws[f'O{row_count}'] = total_dict.get('pt')
	ws[f'P{row_count}'] = total_dict.get('net_amount_paid')

	for row in ws.iter_rows():
		for cell in row:
			cell.border = thin_border

	ws.freeze_panes = 'E6'
	# Save the workbook in memory
	excel_file = BytesIO()
	wb.save(excel_file)
	excel_file.seek(0)
	excel_base64 = base64.b64encode(excel_file.getvalue()).decode('utf-8')

	# Return the file as a response for direct download
	return {
		'file_name': f"FORM XVII-{full_month_name} {year}.xlsx",
		'file_content': excel_base64
	}